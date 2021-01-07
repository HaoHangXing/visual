import openpyxl
from openpyxl.styles import Font, colors, Alignment
from openpyxl.utils import get_column_letter

from openpyxl.chart import ScatterChart, Series, Reference, LineChart
from openpyxl.chart.layout import Layout, ManualLayout



class StrTitle(object):
    # 数据格式
    title_offset = 19 # 开头格式对齐
    s_run_time = '[time]'
    s_dc_vol = '[dc vol]'
    s_ch_vol = '[ch_vol]'
    s_test_ch = '[channel]'
    s_test_index = '[index]'
    s_cell_vol = u'通道电压(V)'
    s_test_vol = u'开关后电压(V)'
    s_rad_tem = u'散热片温度(℃)'
    s_board_tem = u'电路板温度(℃)'
    s_current = u'电流(A)'

    s_title = u'参数'
    s_ch = u'通道'

class SheetNameIndex(object):
    '''
    一些固定名称的sheet
    '''
    t_main = 'main'
    t_total = 'total'
    sheet_title = [t_main, t_total]


class CellAlignForamt(object):
    align_center = Alignment(horizontal='center', vertical='center')


class BaseSheetInfo(CellAlignForamt):
    ws = ''
    name = ''
    index = ''

    w_line = 0 
    s_x = 1
    s_y = 1
    index = 1

    def __init__(self, ws, name, index=None):
        self.ws = ws
        self.name = name
        if index == None:
            self.index = 0
        else:
            self.index = index


    def InsertRows(self, row, num):
        for i in range(0, num):
            self.ws.insert_rows(self.w_y+i)

    def WriteRowList(self, colume, row, list):
        for i, var in enumerate(list):
            self.ws.cell(row, colume+i, var)
            self.ws.cell(row, colume+i).alignment = self.align_center
            

    def SetRangeAlignment(self, s_col, s_row, e_col, e_row, alignment):
        for y in range(s_row, e_row):
            for x in range(s_col, e_col):
                self.ws.cell(y, x).alignment = alignment


# 记录total数据的sheet
class WsTotal(StrTitle, BaseSheetInfo, CellAlignForamt, SheetNameIndex):
    sheet_index = 0

    def __init__(self, wb, name):
        ws = wb.create_sheet(name)
        BaseSheetInfo.__init__(self, ws, self.t_total, self.sheet_title.index(self.t_total))


    def InitSheetBorder(self):
        self.ws.column_dimensions['A'].width = 15
        self.ws['A1'] = self.s_title
        self.ws['B1'] = self.s_ch
        self.ws['B1'].alignment = self.align_center
        self.w_line += 1

       
    def WriteData(self, data):
        # self.ws_dict[self.total].s_x += 1
        self.WriteOneLine(self.s_cell_vol,  data.ch,  data.cell_vol_list)
        self.WriteOneLine(self.s_test_vol,  data.ch,  data.test_vol_list)
        self.WriteOneLine(self.s_rad_tem,   data.ch,  data.rad_temperature_list)
        self.WriteOneLine(self.s_board_tem, data.ch,  data.board_temperature_list)
        self.WriteOneLine(self.s_current,   data.ch,  data.test_current_list)
        # self.WriteEmptyLine(self.total, 2)
  

    def WriteOneLine(self, title, ch, data_list):
        pass
        y = self.s_y + self.w_line
        x = self.s_x
        
        self.ws.cell(y, x, value = title)
        x += 1
        self.ws.cell(y, x, value = ch)
        self.ws.cell(y,x).alignment = self.align_center
        x += 1
        self.WriteRowList(x, y, data_list)
        self.SetRangeAlignment(x, y, x+len(data_list), y, self.align_center)
        self.w_line += 1


    def WriteEmptyLine(self, ws_name, num):
        self.w_line += num
        
# 记录通道数据的sheet， 因为和总表很像，用到相同的函数，就直接继承使用了
class WsChannel(WsTotal):
    def __init__(self, wb, ch_title, ch):
        ws = wb.create_sheet(ch_title)
        BaseSheetInfo.__init__(self, ws, ch_title, ch)
        self.ch = ch

        
    def InitSheetBorder(self):
        self.ws.column_dimensions['A'].width = 15
        self.ws['A1'] = self.s_title
        self.w_line += 1

        for i in range(1, 361):
            cell = "%s%d" % (get_column_letter(i+self.s_x), self.s_y)
            self.ws[cell] = i

    def WriteData(self, data):
        self.WriteOneLine(self.s_cell_vol,  data.cell_vol_list)
        self.WriteOneLine(self.s_test_vol,  data.test_vol_list)
        chart = LineCharts(self.ws)
        # chart.SetChartsXValue(self.s_x+1, self.s_y, self.s_x+len(data.test_vol_list))
        # chart.SetChartsYValue(self.s_x, self.w_line, self.s_x+len(data.test_vol_list))
        chart.SetChartsXValue(self.s_x+1, self.s_y, self.s_x+10)
        chart.SetChartsYValue(self.s_x, self.w_line, self.s_x+10)
        GetSheetLineData(self.ws, self.s_x, self.w_line, self.s_x+10, self.w_line)
        
        self.WriteOneLine(self.s_rad_tem,   data.rad_temperature_list)
        self.WriteOneLine(self.s_board_tem, data.board_temperature_list)
        self.WriteOneLine(self.s_current,   data.test_current_list)

        chart.SetChartsTitle()
        cell = "%s%d" % (get_column_letter(self.s_x+1), self.w_line+2)
        chart.DrawCell(cell)
        self.WriteEmptyLine(self.ws, 16)
    
    def WriteOneLine(self, title, data_list):
        pass
        y = self.s_y + self.w_line
        x = self.s_x
        
        self.ws.cell(y, x, value = title)
        x += 1

        self.WriteRowList(x, y, data_list)
        self.SetRangeAlignment(x, y, x+len(data_list), y, self.align_center)
        
        self.w_line += 1

# 从总表中获取想要的数据
class WsMainData():
    def __init__(self, ch):
        self.ch = ch
        self.num = 2

    def __add__(self, other):
        c_add = WsMainData()
        c_add.s_y = self.s_y + self.num + other.num
        return c_add

class WsMain(BaseSheetInfo, SheetNameIndex):
    def __init__(self, wb):
        ws = wb.create_sheet(self.t_main)
        BaseSheetInfo.__init__(self, ws, self.t_main, self.sheet_title.index(self.t_main))

    def SetReadSheet(self, ws):
        self.r_ws = ws

    def MainHandler(self):
        r_x = 2
        r_y = 2
        c_list = []
        # 提取最后5个通道电压
        cell_vol_offset = 197-60
        cell_vol_num = 4
        # 提取160个开关电压
        test_vol_offset = 1
        test_vol_num = 159


        # 写坐标
        self.w_x = 1
        self.w_y = 1
        # 
        end_empty_line = 3
        while True:
            self.w_y = 1
            data = self.r_ws.cell(row=r_y, column=r_x)
            ch = data.value

            if data.value == None:
                break
 
            for i, ch_data in enumerate(c_list):
                self.w_y += ch_data.num+end_empty_line
                if ch_data.ch == ch:
                    ch_data.num += 1
                    self.w_y -= end_empty_line
                    break;
                elif ch_data.ch > ch:
                    ch_data = WsMainData(ch)
                    c_list.insert(i, ch_data)
                    self.w_y -= ch_data.num + end_empty_line
                    self.AppendTitleLine(cell_vol_num, test_vol_num)
                    self.AppendEndLine(ch)
                    break
                
            else:
                ch_data = WsMainData(ch)
                c_list.append(ch_data)
                self.AppendTitleLine(cell_vol_num, test_vol_num)
                self.AppendEndLine(ch)

            cell_vol = GetSheetLineData(self.r_ws, r_x+cell_vol_offset, r_y, r_x+cell_vol_offset+cell_vol_num, r_y) 
            test_vol = GetSheetLineData(self.r_ws, r_x+test_vol_offset, r_y+1, r_x+test_vol_offset+test_vol_num, r_y+1)

            tmp_list = [ch, ch_data.num-1]
            tmp_list += cell_vol
            tmp_list.append('')
            tmp_list += test_vol

            
            self.InsertRows(self.w_y, 1)
            for i, value in enumerate(tmp_list):
                self.ws.cell(column=self.w_x+i, row=self.w_y, value =value)

            r_y += 5
            #for c in c_list:
            #    print("ch:%d num:%d"%(c.ch, c.num))
            #print("data ch:%d num:%d self.w_y:%d"%(ch_data.ch, ch_data.num, self.w_y))
            #print("====================")
        
    def AppendTitleLine(self, cell_vol_num, test_vol_num):
        tmp_list = ['通道','序号']
        s_num = 140 # 200-60
        tmp_list += [x for x in range(s_num-cell_vol_num, s_num+1)]
        tmp_list.append('')
        tmp_list += [x for x in range(s_num+1, s_num+1+test_vol_num+1)]
        self.InsertRows(self.w_y, 1)
        self.WriteRowList(self.w_x, self.w_y, tmp_list)
        self.w_y += 1

    def AppendEndLine(self, ch):
        #res_list = [1.246,1.111,3.31,1.913,1.882,1.754,1.243,1.487,1.318,1.428,
        #            1.797,1.619,0.771,1.638,1.024,1.198,0.797,1.985,1.415,1.651,
        #            0.864,0.91,0.984,1.045]
        #res_list = [1.54,1.127,3.43,1.915,1.871,1.778,1.23,1.489,1.323,1.362,
        #            1.768,1.666,0.789,1.648,1.038,1.204,0.78,2.002,1.378,1.631,
        #            0.884,0.928,0.988,1.386]
        res_list = [1.27  ,1.121 ,3.42 ,1.94  ,1.87  ,1.773 ,1.224 ,1.499 ,1.349 ,1.343 ,
                    1.774 ,1.889 ,0.804 ,1.65  ,1.044 ,1.199 ,0.78  ,1.983 ,1.385 ,1.623 ,
                    0.899 ,0.959 ,0.988 ,1.351]
        tmp_list = ['实际内阻']
        tmp_list.append(res_list[ch])
        

        self.InsertRows(self.w_y, 3)
        self.WriteRowList(self.w_x, self.w_y, tmp_list)

class DCharts(object):
    def __init__(self, ws):
        self.ws = ws
        self.ch1 = LineChart() # 创建图s
        
    def SetChartsXValue(self, s_x=None, s_y=None, e_x=None, e_y=None):
        self.valuesx = Reference(self.ws, s_x, s_y, e_x, e_y)

    def SetChartsYValue(self, s_x=None, s_y=None, e_x=None, e_y=None):
        values = Reference(self.ws, s_x, s_y, e_x, e_y)
        series = Series(values, self.valuesx, title_from_data=True)
        self.ch1.series.append(series)
        
        #xvalues = Reference(ws, min_col=1, min_row=2, max_row=7)
        #for i in range(2, 4):
        #    values = Reference(ws, min_col=i, min_row=1, max_row=7)
        #    series = Series(values, xvalues, title_from_data=True)
        #    ch1.series.append(series)
    
    def SetChartsTitle(self):
        self.ch1.title = "前10个开关电压"
        self.ch1.style = 10
        self.ch1.x_axis.title = 'Vol'
        self.ch1.y_axis.title = 'Percentage'
        self.ch1.legend.position = 'r' # 图例位置

        s1 = self.ch1.series[0]
        s1.marker.symbol = "triangle"
        s1.marker.graphicalProperties.solidFill = "FF0000" # Marker filling
        s1.marker.graphicalProperties.line.solidFill = "FF0000" # Marker outline

        #s1.graphicalProperties.line.noFill = True

        #s2 = self.ch1.series[1]
        #s2.graphicalProperties.line.solidFill = "00AAAA"
        #s2.graphicalProperties.line.dashStyle = "sysDot"
        #s2.graphicalProperties.line.width = 100050 # width in EMUs

        #s2 = self.ch1.series[2]
        #s2.smooth = True # Make the line smooth

    def DrawCell(self, cell):
        self.ws.add_chart(self.ch1, cell)


# 用于画图的类
class LineCharts(object):
    def __init__(self, ws):
        self.ws = ws
        self.ch1 = ScatterChart() # 创建图s
        
    def SetChartsXValue(self, s_x=None, s_y=None, e_x=None, e_y=None):
        self.valuesx = Reference(self.ws, s_x, s_y, e_x, e_y)

    def SetChartsYValue(self, s_x=None, s_y=None, e_x=None, e_y=None):
        values = Reference(self.ws, s_x, s_y, e_x, e_y)
        series = Series(values, self.valuesx, title_from_data=True)
        self.ch1.series.append(series)
        
        #xvalues = Reference(ws, min_col=1, min_row=2, max_row=7)
        #for i in range(2, 4):
        #    values = Reference(ws, min_col=i, min_row=1, max_row=7)
        #    series = Series(values, xvalues, title_from_data=True)
        #    ch1.series.append(series)
    
    def SetChartsTitle(self):
        self.ch1.title = "前10个开关电压"
        self.ch1.style = 10
        self.ch1.x_axis.title = 'Vol'
        self.ch1.y_axis.title = 'Percentage'
        self.ch1.legend.position = 'r' # 图例位置
    #    self.ch1.layout=Layout(
    #    manualLayout=ManualLayout(
    #        # x=0.25, y=0.25,
    #        h=0.5, w=0.5,
    #    )
    #)


    def DrawCell(self, cell):
        self.ws.add_chart(self.ch1, cell)

class ELog(SheetNameIndex):
    def __init__(self, file):
        self.wb = openpyxl.Workbook()
        print("set write to  :"+ file)
        self.save = file

        self.ws_dict = {} # 存储通道sheet类
        
        # 创建一个总表
        self.total_ws = WsTotal(self.wb, self.t_total)
        # self.ws_dict[self.total_ws.name] = self.total_ws 
        self.total_ws.InitSheetBorder()
     
          
    def close(self):
        self.wb.save(self.save)
        self.wb.close()


    def InputWData(self, c_data):
        self.data = c_data


    def WriteOutLog(self):
        # 每个通道单独创建sheet存储
        # 再创建一个总sheet存储所有通道
        ws_name = f'ch{self.data.ch}'

        if ws_name in self.ws_dict.keys():
            ch_ws = self.ws_dict[ws_name]
        else:
            ch_ws = WsChannel(self.wb, ws_name, self.data.ch)
            ch_ws.InitSheetBorder()
            self.ws_dict[ch_ws.name] = ch_ws

        ch_ws.WriteData(self.data)
        self.total_ws.WriteData(self.data)
    
    def WsMainOrganizeData(self):
        self.main_ws = WsMain(self.wb)
        self.main_ws.SetReadSheet(self.total_ws.ws)
        self.main_ws.MainHandler()


    def SortSheet(self):
        tmp_list = sorted(self.ws_dict.values(), key=lambda x : x.ch) # 排序通道sheet
        for i, sheet in enumerate(tmp_list):
            self.wb.move_sheet_by_index(sheet.name, i)

        for i, sheet in enumerate(self.sheet_title):
            self.wb.move_sheet_by_index(sheet, i)
        


def GetSheetLineData(ws, s_x, s_y, e_x, e_y):
    cell_range = ws.iter_rows(min_col=s_x, min_row=s_y, max_col=e_x, max_row=e_y)
    tmp_list = []
    for row in cell_range:
        for cell in row:
            tmp_list.append(cell.value)
    return tmp_list

