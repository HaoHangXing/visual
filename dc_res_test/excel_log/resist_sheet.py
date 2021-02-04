from excel import base as excel_base
from openpyxl.utils import get_column_letter
from excel_log.total_sheet import TotalInfo

class WsResistData():
    def __init__(self, ch):
        self.ch = ch
        self.num = 1

    def __add__(self, other):
        c_add = WsMainData()
        c_add.s_y = self.s_y + self.num + other.num
        return c_add

# resist sheet 从total表中获取想要的数据
class WsResist(excel_base.SheetInfo):
    def __init__(self, wb, name, index):
        ws = wb.create_sheet(name)
        excel_base.SheetInfo.__init__(self, ws, name, index)
        self.s_x = 1
        self.s_y = 3

    def SetReadSheet(self, ws):
        self.r_ws = ws

    def MainHandler(self):
        read_info = TotalInfo()
        r_x = read_info.r_x
        r_y = read_info.r_y
        cell_vol_line = read_info.line_info.index(read_info.cell_vol_line)
        test_vol_line = read_info.line_info.index(read_info.test_vol_line)
        current_line = read_info.line_info.index(read_info.current_line)
        c_list:channel_class = []
        
        self.title_line = 2  # 标题占2行
        self.channel_row = 4 # 一个通道提供4列数据
        
        self.max_num = 0
        while True:
            self.w_x = self.s_x + 1
            self.w_y = self.s_y
            data = self.r_ws.cell(row=r_y, column=r_x)
            ch = data.value
            if data.value == None:
                break
 
            for i, ch_data in enumerate(c_list):
                if ch_data.ch == ch:
                    self.w_y += self.title_line + ch_data.num
                    ch_data.num += 1
                    break;
                elif ch_data.ch > ch:
                    ch_data = WsResistData(ch)
                    c_list.insert(i, ch_data)
                    self.AppendTitleLRow(ch)
                    break
                self.w_x += self.channel_row
            else:
                ch_data = WsResistData(ch)
                c_list.append(ch_data)
                self.AppendTitleLRow(ch)
                #self.AppendTitleLRow(cell_vol_num, test_vol_num)
                #self.AppendEndLine(ch)

            if self.max_num < ch_data.num:
                self.max_num = ch_data.num

            # 填写数据
            if 1:
                cell = self.r_ws.cell(row=r_y+test_vol_line, column=r_x+1)
                vol = cell.value
            else:
                cell1 = self.r_ws.cell(row=r_y+cell_vol_line, column=r_x+1+self.data.except_cell_vol_num1+self.data.except_cell_vol_num2)
                cell2 = self.r_ws.cell(row=r_y+test_vol_line, column=r_x+1)
                vol = cell2.value - cell1.value

            cell = self.r_ws.cell(row=r_y+current_line, column=r_x+1)
            current = cell.value
            resist = "=%s%d/%s%d/220*1000" % (get_column_letter(self.w_x), self.w_y, get_column_letter(self.w_x+1), self.w_y) 
            #resist = "=%s%d/%s%d/154.5*1000" % (get_column_letter(self.w_x), self.w_y, get_column_letter(self.w_x+1), self.w_y) 
            #resist = "=%s%d/%s%d*1000" % (get_column_letter(self.w_x), self.w_y, get_column_letter(self.w_x+1), self.w_y) 
            data_list = [vol, current, resist]
            self.WriteRowList(self.w_x, self.w_y, data_list)

            r_y += len(read_info.line_info)
        # 填写左侧标题栏
        self.AppendLeftTitle()
        
        # 填写结尾数据
        self.AppendEndLine(c_list)


    def AppendTitleLRow(self, ch):
        title_list = ['第一次开关电压(V)','电流(A)','电阻(mΩ)','加权电阻(mΩ)']
        len_list = len(title_list)
        self.InsertCols(self.w_x, len_list)
        self.ws.cell(self.w_y, self.w_x, value = f'通道{ch}')
        self.w_y+=1
        self.WriteRowList(self.w_x, self.w_y, title_list)
        self.w_y+=1

    def AppendLeftTitle(self):
        w_x = self.s_x
        w_y = self.s_y+1
        self.ws.cell(w_y, w_x, value = '测量次数')
        num_list = [int(x) for x in range(1,self.max_num+1)]
        self.WriteColsList(w_x, w_y+1, num_list)


    def AppendEndLine(self, ch_list):
        # 1楼24节220Ah2V
        #res_list = [1.246,1.111,3.31,1.913,1.882,1.754,1.243,1.487,1.318,1.428,
        #            1.797,1.619,0.771,1.638,1.024,1.198,0.797,1.985,1.415,1.651,
        #            0.864,0.91,0.984,1.045]
        #res_list = [1.54,1.127,3.43,1.915,1.871,1.778,1.23,1.489,1.323,1.362,
        #            1.768,1.666,0.789,1.648,1.038,1.204,0.78,2.002,1.378,1.631,
        #            0.884,0.928,0.988,1.386]
        #res_list = [1.27  ,1.121 ,3.42 ,1.94  ,1.87  ,1.773 ,1.224 ,1.499 ,1.349 ,1.343 ,
        #            1.774 ,1.889 ,1.804 ,1.65  ,1.044 ,1.199 ,0.78  ,1.983 ,1.385 ,1.623 ,
        #            0.899 ,0.959 ,0.988 ,1.351]
        if 0:
            # 2楼24节200Ah2V
            res_list = [0.63  ,0.634 ,0.616 ,0.629 ,0.608 ,0.625 ,0.628 ,0.609 ,0.598 ,0.646 ,
                        0.632 ,0.602 ,0.629 ,0.644 ,0.602 ,0.621 ,0.637 ,0.606 ,0.583 ,0.573 ,0.621 ,
                        0.62  ,0.603 ,0.598 ,]
        elif 0:
            # 2楼4节100Ah12V
            res_list = [5.2, 6.11, 5.78, 5.12]
        elif 1:
            res_list = [8.34,8.46,8.4 ,8.34,8.08,8.31,8.5 ,8.45,8.33,8.34,
            8.29,8.32,8.25,8.36,8.15,8.3 ,8.42,8.42,8.37,8.1 ,
            8.44,8.25,8.14,8.38,8.4 ,8.82,8.38,8.21,8.16,8.32,]

        title_list = ['最大值','最小值','极差','实际阻值','正偏','负偏']
        w_x = self.s_x
        w_y = self.s_y+self.title_line+self.max_num
        self.WriteColsList(w_x, w_y, title_list)

        w_x += 1
        for i, ch_data in enumerate(ch_list):
            x_offset = w_x + i*self.channel_row
            for j in range(0,self.channel_row):
                self.ws.cell(w_y,   x_offset+j, value = '=MAX(%s%d:%s%d)'%(get_column_letter(x_offset+j), w_y-self.max_num,get_column_letter(x_offset+j),w_y-1))
                self.ws.cell(w_y+1, x_offset+j, value = '=MIN(%s%d:%s%d)'%(get_column_letter(x_offset+j), w_y-self.max_num,get_column_letter(x_offset+j),w_y-1))
                self.ws.cell(w_y+2, x_offset+j, value = '=%s%d-%s%d'     %(get_column_letter(x_offset+j), w_y, get_column_letter(x_offset+j), w_y+1))
            self.ws.cell(w_y+3, x_offset+2, value = res_list[ch_data.ch])
            self.ws.cell(w_y+4, x_offset+2, value = '=%s%d-%s%d'     %(get_column_letter(x_offset+2), w_y, get_column_letter(x_offset+2), w_y+3))
            self.ws.cell(w_y+5, x_offset+2, value = '=%s%d-%s%d'     %(get_column_letter(x_offset+2), w_y+1, get_column_letter(x_offset+2), w_y+3))

    def Inputdata(self,data):
        self.data = data