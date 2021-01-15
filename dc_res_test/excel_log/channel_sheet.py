
from openpyxl.utils import get_column_letter
from excel import base as excel_base
from excel import chart as excel_chart
from excel_log import total_sheet
# 记录通道数据的sheet， 因为和总表很像，用到相同的函数，就直接继承使用了
class WsChannel(total_sheet.WsTotal):
    def __init__(self, wb, ch_title, ch):
        ws = wb.create_sheet(ch_title)
        excel_base.SheetInfo.__init__(self, ws, ch_title, ch)
        self.ch = ch

        
    def InitSheetBorder(self):
        self.ws.column_dimensions['A'].width = 15
        self.ws['A1'] = self.s_title
        self.w_line += 1

        for i in range(1, 361):
            cell = "%s%d" % (get_column_letter(i+self.s_x), self.s_y)
            self.ws[cell] = i

    def WriteData(self, data):
        self.WriteOneLine(self.s_cell_vol,  data.cell_vol_list)
        self.WriteOneLine(self.s_test_vol,  data.test_vol_list)
        chart = excel_chart.LineCharts(self.ws)
        # chart.SetChartsXValue(self.s_x+1, self.s_y, self.s_x+len(data.test_vol_list))
        # chart.SetChartsYValue(self.s_x, self.w_line, self.s_x+len(data.test_vol_list))
        chart.SetChartsXValue(self.s_x+1, self.s_y, self.s_x+10)
        chart.SetChartsYValue(self.s_x, self.w_line, self.s_x+10)
        excel_base.GetSheetLineData(self.ws, self.s_x, self.w_line, self.s_x+10, self.w_line)
        
        self.WriteOneLine(self.s_rad_tem,   data.rad_temperature_list)
        self.WriteOneLine(self.s_board_tem, data.board_temperature_list)
        self.WriteOneLine(self.s_current,   data.test_current_list)

        chart.SetChartsTitle("前10个开关电压", 'Vol', 'Percentage')
        cell = "%s%d" % (get_column_letter(self.s_x+1), self.w_line+2)
        chart.DrawCell(cell)
        self.WriteEmptyLine(self.ws, 16)
    
    def WriteOneLine(self, title, data_list):
        pass
        y = self.s_y + self.w_line
        x = self.s_x
        
        self.ws.cell(y, x, value = title)
        x += 1

        self.WriteRowList(x, y, data_list)
        self.SetRangeAlignment(x, y, x+len(data_list), y, self.align_center)
        
        self.w_line += 1