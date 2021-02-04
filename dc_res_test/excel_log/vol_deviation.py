import tool
from excel import base as excel_base
from excel import chart as excel_chart
from excel_log.total_sheet import TotalInfo
from openpyxl.utils import get_column_letter

class WsMainData():
    def __init__(self, ch):
        self.ch = ch
        self.num = 1

    def __add__(self, other):
        c_add = WsMainData()
        c_add.s_y = self.s_y + self.num + other.num
        return c_add
class WsVolDevi(excel_base.SheetInfo):
    def __init__(self, wb, name, index):
        ws = wb.create_sheet(name)
        excel_base.SheetInfo.__init__(self, ws, name, index)

    def SetReadSheet(self, ws):
        self.r_ws = ws

    def MainHandler(self):
        read_info = TotalInfo()
        r_x = read_info.r_x
        r_y = read_info.r_y
        c_list = []
        still_data_flag = True
        chart_line = 21 # 留给画图的行数

        # 提取160个开关电压
        test_vol_offset = 1
        test_vol_num = self.data.except_test_vol_num - 1
        test_vol_line = read_info.line_info.index(read_info.test_vol_line)

        # 写坐标
        self.w_x = 1
        self.w_y = 1
        # 数据前后的空行，用于写标题和结尾
        self.title_empty_line = 1 
        self.end_empty_line = 5 + chart_line  # 5行数据统计，画图位置
        
        while still_data_flag:
            self.w_y = 1
            data = self.r_ws.cell(row=r_y, column=r_x)
            ch = data.value

            if data.value == None:
                still_data_flag = False
                continue
 
            for i, ch_data in enumerate(c_list):
                self.w_y += ch_data.num + self.title_empty_line + self.end_empty_line
                if ch_data.ch == ch:
                    ch_data.num += 1
                    self.w_y -= self.end_empty_line
                    break;
                elif ch_data.ch > ch:
                    ch_data = WsMainData(ch)
                    c_list.insert(i, ch_data)
                    self.w_y -= ch_data.num + self.title_empty_line + self.end_empty_line
                    self.AppendTitleLine(test_vol_num)
                    self.AppendEmptyLine()
                    break
                
            else:
                ch_data = WsMainData(ch)
                c_list.append(ch_data)
                self.AppendTitleLine(test_vol_num)
                self.AppendEmptyLine()

            #cell_vol = excel_base.GetSheetLineData(self.r_ws, r_x+cell_vol_offset, r_y+cell_vol_line, r_x+cell_vol_offset+cell_vol_num, r_y+cell_vol_line) 
            test_vol = excel_base.GetSheetLineData(self.r_ws, r_x+test_vol_offset, r_y+test_vol_line, r_x+test_vol_offset+test_vol_num, r_y+test_vol_line)

            tmp_list = [ch, ch_data.num]
            #tmp_list += cell_vol
            #tmp_list.append('')
            tmp_list += test_vol

            
            self.InsertRows(self.w_y, 1)
            for i, value in enumerate(tmp_list):
                self.ws.cell(column=self.w_x+i, row=self.w_y, value =value)

            r_y += len(read_info.line_info)
        else:
            self.w_x = 2
            self.w_y = 1
            
            for ch_data in c_list:
                self.w_y += self.title_empty_line + ch_data.num
                self.AppendEndLine()
                #resist = "=%s%d/%s%d/220*1000" % (get_column_letter(self.w_x), self.w_y, get_column_letter(self.w_x+1), self.w_y) 
                for i in range(1,test_vol_num+1):
                    self.ws.cell(self.w_y,   self.w_x+i, value = '=MAX(%s%d:%s%d)'%(get_column_letter(self.w_x+i), self.w_y-ch_data.num,get_column_letter(self.w_x+i),self.w_y-1))
                    self.ws.cell(self.w_y+1, self.w_x+i, value = '=MIN(%s%d:%s%d)'%(get_column_letter(self.w_x+i), self.w_y-ch_data.num,get_column_letter(self.w_x+i),self.w_y-1))
                    self.ws.cell(self.w_y+2, self.w_x+i, value = '=%s%d-%s%d'     %(get_column_letter(self.w_x+i), self.w_y, get_column_letter(self.w_x+i), self.w_y+1))
                    self.ws.cell(self.w_y+3, self.w_x+i, value = '=AVERAGE(%s%d:%s%d)'%(get_column_letter(self.w_x+i), self.w_y-ch_data.num,get_column_letter(self.w_x+i),self.w_y-1))
                    self.ws.cell(self.w_y+4, self.w_x+i, value = '=%s%d/%s%d'     %(get_column_letter(self.w_x+i), self.w_y+2, get_column_letter(self.w_x+i), self.w_y+3))
                # 画重复度的散点图
                chart = excel_chart.MyScatterChart(self.ws)
                chart.SetChartsTitle('开关电压重复度', '次数（20ms一次）','重复度')
                chart.SetChartWH(25,10)
                chart.SetChartsAddData(self.w_x+1, self.w_y+4, self.w_x+60, self.w_y+4)
                cell = "%s%d" % (get_column_letter(self.w_x+1), self.w_y+5)
                chart.DrawCell(cell)
                self.w_y += self.end_empty_line

    def AppendTitleLine(self, test_vol_num):
        tmp_list = ['通道','序号']
        start = 1
        tmp_list += [x for x in range(start, start+test_vol_num)]
        self.InsertRows(self.w_y, 1)
        self.WriteRowList(self.w_x, self.w_y, tmp_list)
        self.w_y += self.title_empty_line

    def AppendEndLine(self):
        tmp_list = ['最大值','最小值','极差','平均值','重复度']
        self.WriteColsList(self.w_x, self.w_y, tmp_list)
        

    # 预留空行给数据统计和画图
    def AppendEmptyLine(self):
        self.InsertRows(self.w_y, self.end_empty_line)


    def Inputdata(self,data):
        self.data = data
