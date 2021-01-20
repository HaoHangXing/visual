import openpyxl
from openpyxl.styles import Font, colors, Alignment


class CellAlignForamt(object):
    align_center = Alignment(horizontal='center', vertical='center')

class CellLocate(object):
    x = 0
    y = 0
    


class SheetInfo(CellAlignForamt):
    ws = ''
    name = ''
    index = ''

    s_x = 1
    s_y = 1
    index = 1

    def __init__(self, ws, name, index=None):
        self.ws = ws
        self.name = name
        if index == None:
            self.index = 0
        else:
            self.index = index


    def InsertRows(self, row, num):
        for i in range(0, num):
            self.ws.insert_rows(row+i)

    def InsertCols(self, col, num):
        for i in range(0, num):
            self.ws.insert_cols(col+i)

    def WriteRowList(self, colume, row, list):
        for i, var in enumerate(list):
            self.ws.cell(row, colume+i, var)
            self.ws.cell(row, colume+i).alignment = self.align_center

    def WriteColsList(self, colume, row, list):
        for i, var in enumerate(list):
            self.ws.cell(row+i, colume, var)
            self.ws.cell(row+i, colume).alignment = self.align_center
            

    def SetRangeAlignment(self, s_col, s_row, e_col, e_row, alignment):
        for y in range(s_row, e_row):
            for x in range(s_col, e_col):
                self.ws.cell(y, x).alignment = alignment


def GetSheetLineData(ws, s_x, s_y, e_x, e_y):
    cell_range = ws.iter_rows(min_col=s_x, min_row=s_y, max_col=e_x, max_row=e_y)
    tmp_list = []
    for row in cell_range:
        for cell in row:
            tmp_list.append(cell.value)
    return tmp_list